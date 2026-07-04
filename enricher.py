"""
AWB-level charge enrichment.

Derives GST-inclusive handling, X-ray and demurrage charges from
airline service sheets and reconciles rounding residuals back to the
source Service Challan total.
"""

from pathlib import Path

import openpyxl
import pandas as pd

from adapters.bz_adapter import read_bz_summary
from config import GST_RATE
from reconciler import read_8m_sc_totals, read_standard_sc_total
from utils import clean_text, extract_airline, safe_numeric


CHARGE_COLUMNS = [
    "Handling Amt (Incl GST)",
    "X-Ray Amt (Incl GST)",
    "Demurrage Amt (Incl GST)",
]


def normalise_awb(value):
    """Create a stable AWB lookup key."""
    return clean_text(value).upper().replace(" ", "")


def find_header_row(rows, required_words):
    """Find a row containing one of the required header words."""
    for index, row in enumerate(rows):
        values = [
            clean_text(value).upper()
            for value in row
        ]

        joined = " | ".join(values)

        if all(
            any(word in value for value in values)
            for word in required_words
        ):
            return index

    return None


def find_column(headers, candidates):
    """Return the index of the first matching header."""
    normalised = [
        clean_text(value).upper()
        for value in headers
    ]

    for candidate in candidates:
        candidate = candidate.upper()

        for index, header in enumerate(normalised):
            if candidate in header:
                return index

    return None


def read_sheet_rows(file_path, sheet_name):
    """Read worksheet values into a list."""
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )

    worksheet = workbook[sheet_name]

    return list(
        worksheet.iter_rows(values_only=True)
    )


def extract_amount_lookup(
    file_path,
    sheet_name,
    amount_candidates,
):
    """
    Read AWB -> GST-inclusive amount from a service detail sheet.
    """
    rows = read_sheet_rows(
        file_path,
        sheet_name,
    )

    header_index = find_header_row(
        rows,
        required_words=["AWB"],
    )

    if header_index is None:
        return {}

    headers = rows[header_index]

    awb_index = find_column(
        headers,
        ["AWB.NO", "AWB NO", "AWB", "MAWB"],
    )

    amount_index = find_column(
        headers,
        amount_candidates,
    )

    if awb_index is None or amount_index is None:
        return {}

    lookup = {}

    for row in rows[header_index + 1:]:
        if awb_index >= len(row):
            continue

        awb = normalise_awb(row[awb_index])

        if not awb or "TOTAL" in awb:
            continue

        if amount_index >= len(row):
            continue

        amount = safe_numeric(
            row[amount_index]
        )

        lookup[awb] = (
            lookup.get(awb, 0.0)
            + amount
        )

    return lookup


def get_demurrage_lookup(file_path):
    """Read GST-inclusive demurrage values."""
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )

    candidate_sheets = [
        name
        for name in workbook.sheetnames
        if (
            "DEMURRAGE" in name.upper()
            or name.upper().startswith("DEM")
        )
    ]

    combined = {}

    for sheet_name in candidate_sheets:
        lookup = extract_amount_lookup(
            file_path,
            sheet_name,
            [
                "AMT INCLUDING GST",
                "AMOUNT INCLUDING GST",
                "AMT INCL GST",
                "AMOUNT INCL GST",
                "TOTAL AMOUNT",
            ],
        )

        for awb, amount in lookup.items():
            combined[awb] = (
                combined.get(awb, 0.0)
                + amount
            )

    return combined


def get_direct_charge_lookup(
    file_path,
    sheet_keywords,
):
    """
    Read direct GST-inclusive service amounts when the workbook
    provides an AWB-level amount column.
    """
    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )

    combined = {}

    for sheet_name in workbook.sheetnames:
        upper_name = sheet_name.upper()

        if not any(
            keyword in upper_name
            for keyword in sheet_keywords
        ):
            continue

        lookup = extract_amount_lookup(
            file_path,
            sheet_name,
            [
                "AMT INCLUDING GST",
                "AMOUNT INCLUDING GST",
                "AMT INCL GST",
                "AMOUNT INCL GST",
                "TOTAL AMOUNT",
            ],
        )

        for awb, amount in lookup.items():
            combined[awb] = (
                combined.get(awb, 0.0)
                + amount
            )

    return combined


def allocate_amount_by_weight(
    dataframe,
    row_indexes,
    amount,
    output_column,
):
    """
    Allocate a source service amount proportionally by chargeable
    weight. This is used only when the workbook exposes a service
    total but not a direct AWB amount.
    """
    if not row_indexes or amount == 0:
        return

    weights = dataframe.loc[
        row_indexes,
        "Chg Wt",
    ].astype(float)

    total_weight = weights.sum()

    if total_weight > 0:
        allocated = (
            weights / total_weight
        ) * amount

    else:
        allocated = pd.Series(
            amount / len(row_indexes),
            index=row_indexes,
        )

    dataframe.loc[
        row_indexes,
        output_column,
    ] += allocated


def source_target_total(file_path, fortnight=None):
    """Read the source invoice total."""
    airline = extract_airline(file_path)

    if airline == "8M":
        totals = read_8m_sc_totals(
            file_path
        )

        return safe_numeric(
            totals.get(fortnight, 0.0)
        )

    if airline == "BZ":
        summary = read_bz_summary(
            file_path
        )

        return safe_numeric(
            summary.get("grand_total", 0.0)
        )

    return safe_numeric(
        read_standard_sc_total(file_path)
    )


def enrich_group(
    dataframe,
    row_indexes,
    file_path,
    fortnight,
):
    """
    Enrich one source-file / fortnight AWB group.
    """
    for column in CHARGE_COLUMNS:
        dataframe.loc[
            row_indexes,
            column,
        ] = 0.0

    demurrage_lookup = get_demurrage_lookup(
        file_path
    )

    handling_lookup = get_direct_charge_lookup(
        file_path,
        [
            "HANDLING",
            "OFFLOADING",
        ],
    )

    xray_lookup = get_direct_charge_lookup(
        file_path,
        [
            "X-RAY",
            "XRAY",
            "SCREENING",
            "CERTIFICATION",
        ],
    )

    for index in row_indexes:
        awb = normalise_awb(
            dataframe.at[index, "AWB No"]
        )

        dataframe.at[
            index,
            "Handling Amt (Incl GST)",
        ] = safe_numeric(
            handling_lookup.get(awb, 0.0)
        )

        dataframe.at[
            index,
            "X-Ray Amt (Incl GST)",
        ] = safe_numeric(
            xray_lookup.get(awb, 0.0)
        )

        dataframe.at[
            index,
            "Demurrage Amt (Incl GST)",
        ] = safe_numeric(
            demurrage_lookup.get(awb, 0.0)
        )

    target_total = source_target_total(
        file_path,
        fortnight,
    )

    current_charge_total = dataframe.loc[
        row_indexes,
        CHARGE_COLUMNS,
    ].sum().sum()

    unallocated_amount = (
        target_total - current_charge_total
    )

    if abs(unallocated_amount) > 0.0001:
        allocate_amount_by_weight(
            dataframe=dataframe,
            row_indexes=row_indexes,
            amount=unallocated_amount,
            output_column="Handling Amt (Incl GST)",
        )

    dataframe.loc[
        row_indexes,
        "Total Amt (Incl GST)",
    ] = dataframe.loc[
        row_indexes,
        CHARGE_COLUMNS,
    ].sum(axis=1)


def enrich_charges(
    consolidated_df,
    source_files,
    skipped_reference_files=None,
):
    """
    Add AWB-level GST-inclusive charge amounts to consolidated data.
    """
    dataframe = consolidated_df.copy()

    skipped_names = {
        Path(path).name
        for path in (
            skipped_reference_files or []
        )
    }

    source_lookup = {
        Path(path).name: Path(path)
        for path in source_files
        if Path(path).name not in skipped_names
    }

    grouped = dataframe.groupby(
        [
            "Source File",
            "Fortnight",
        ],
        sort=False,
    )

    for (
        source_file,
        fortnight,
    ), group in grouped:
        file_path = source_lookup.get(
            source_file
        )

        if file_path is None:
            continue

        enrich_group(
            dataframe=dataframe,
            row_indexes=group.index.tolist(),
            file_path=file_path,
            fortnight=fortnight,
        )

    for column in CHARGE_COLUMNS:
        dataframe[column] = (
            dataframe[column]
            .astype(float)
            .round(4)
        )

    dataframe[
        "Total Amt (Incl GST)"
    ] = (
        dataframe[
            "Total Amt (Incl GST)"
        ]
        .astype(float)
        .round(4)
    )

    return dataframe