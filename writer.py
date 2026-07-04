"""
Final MIS workbook writer.

Copies the supplied template and writes:
- Export AWB-level detail
- Import AWB-level detail
- Month_Summary
- Reconciliation
"""

from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_FILE_NAME, TARGET_COLUMNS


SUMMARY_COLUMNS = [
    "Direction",
    "Airline",
    "Month",
    "AWB Lines",
    "Pcs",
    "Gross Wt",
    "Chg Wt",
    "Handling Amt (Incl GST)",
    "X-Ray Amt (Incl GST)",
    "Demurrage Amt (Incl GST)",
    "Total Amt (Incl GST)",
]


def build_month_summary(dataframe):
    """Create Airline x Month x Direction totals."""
    if dataframe.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    summary = (
        dataframe.groupby(
            ["Direction", "Airline", "Month"],
            as_index=False,
            dropna=False,
        )
        .agg(
            **{
                "AWB Lines": ("AWB No", "size"),
                "Pcs": ("Pcs", "sum"),
                "Gross Wt": ("Gross Wt", "sum"),
                "Chg Wt": ("Chg Wt", "sum"),
                "Handling Amt (Incl GST)": (
                    "Handling Amt (Incl GST)",
                    "sum",
                ),
                "X-Ray Amt (Incl GST)": (
                    "X-Ray Amt (Incl GST)",
                    "sum",
                ),
                "Demurrage Amt (Incl GST)": (
                    "Demurrage Amt (Incl GST)",
                    "sum",
                ),
                "Total Amt (Incl GST)": (
                    "Total Amt (Incl GST)",
                    "sum",
                ),
            }
        )
    )

    amount_columns = [
        "Handling Amt (Incl GST)",
        "X-Ray Amt (Incl GST)",
        "Demurrage Amt (Incl GST)",
        "Total Amt (Incl GST)",
    ]

    for column in amount_columns:
        summary[column] = summary[column].round(4)

    return summary[SUMMARY_COLUMNS]


def clear_sheet(worksheet):
    """Remove all existing rows from a worksheet."""
    if worksheet.max_row > 0:
        worksheet.delete_rows(
            1,
            worksheet.max_row,
        )


def write_dataframe(worksheet, dataframe):
    """Write a DataFrame to an openpyxl worksheet."""
    clear_sheet(worksheet)

    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=column_name,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, row in enumerate(
        dataframe.itertuples(
            index=False,
            name=None,
        ),
        start=2,
    ):
        for column_index, value in enumerate(
            row,
            start=1,
        ):
            if pd.isna(value):
                value = None

            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()

            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    format_worksheet(
        worksheet,
        dataframe.columns.tolist(),
    )


def format_worksheet(worksheet, columns):
    """Apply practical MIS formatting."""
    for index, column_name in enumerate(
        columns,
        start=1,
    ):
        letter = get_column_letter(index)

        max_length = len(str(column_name))

        for cell in worksheet[letter]:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            letter
        ].width = min(
            max(max_length + 2, 12),
            35,
        )

    date_columns = {
        "Flight Date",
    }

    amount_columns = {
        "Gross Wt",
        "Chg Wt",
        "Handling Amt (Incl GST)",
        "X-Ray Amt (Incl GST)",
        "Demurrage Amt (Incl GST)",
        "Total Amt (Incl GST)",
        "Parsed Total",
        "Source SC Total",
        "Difference",
    }

    for column_index, column_name in enumerate(
        columns,
        start=1,
    ):
        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
            )

            if column_name in date_columns:
                cell.number_format = "dd-mmm-yyyy"

            elif column_name in amount_columns:
                cell.number_format = '#,##0.0000'

    worksheet.row_dimensions[1].height = 32


def get_or_create_sheet(workbook, sheet_name):
    """Return an existing sheet or create a new one."""
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    return workbook.create_sheet(
        title=sheet_name
    )


def write_output_workbook(
    base_dir,
    consolidated_df,
    reconciliation_df,
):
    """Generate the final consolidated MIS workbook."""
    base_path = Path(base_dir)

    template_path = (
        base_path
        / "MIS_Consolidated_TEMPLATE.xlsx"
    )

    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}"
        )

    output_dir = base_path / "output"

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path = (
        output_dir
        / OUTPUT_FILE_NAME
    )

    shutil.copy2(
        template_path,
        output_path,
    )

    workbook = load_workbook(
        output_path
    )

    export_df = consolidated_df[
        consolidated_df["Direction"] == "Export"
    ].copy()

    import_df = consolidated_df[
        consolidated_df["Direction"] == "Import"
    ].copy()

    export_df = export_df[
        TARGET_COLUMNS
    ].reset_index(drop=True)

    import_df = import_df[
        TARGET_COLUMNS
    ].reset_index(drop=True)

    month_summary_df = build_month_summary(
        consolidated_df
    )

    export_sheet = get_or_create_sheet(
        workbook,
        "Export",
    )

    import_sheet = get_or_create_sheet(
        workbook,
        "Import",
    )

    summary_sheet = get_or_create_sheet(
        workbook,
        "Month_Summary",
    )

    reconciliation_sheet = get_or_create_sheet(
        workbook,
        "Reconciliation",
    )

    write_dataframe(
        export_sheet,
        export_df,
    )

    write_dataframe(
        import_sheet,
        import_df,
    )

    write_dataframe(
        summary_sheet,
        month_summary_df,
    )

    write_dataframe(
        reconciliation_sheet,
        reconciliation_df,
    )

    workbook.save(
        output_path
    )

    return {
        "output_path": output_path,
        "export_rows": len(export_df),
        "import_rows": len(import_df),
        "summary_rows": len(month_summary_df),
        "reconciliation_rows": len(
            reconciliation_df
        ),
    }