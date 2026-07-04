"""
Source workbook reconciliation.

Reads Service Challan / SC totals from every billing workbook and
compares them with the consolidated AWB charge total for that source.

8M is reconciled separately by fortnight because both fortnights are
stored as separate ANNEX and SC tabs in the same workbook.

BZ uses the SUMMRY grand total.
"""

from pathlib import Path

import openpyxl
import pandas as pd

from adapters.bz_adapter import read_bz_summary
from config import RECONCILIATION_TOLERANCE
from utils import (
    clean_text,
    extract_airline,
    extract_direction,
    extract_fortnight,
    extract_month,
    safe_numeric,
)


def load_workbook_values(file_path):
    """Open workbook using calculated Excel values."""
    return openpyxl.load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )


def find_total_in_sheet(worksheet):
    """
    Find the numeric amount on a row containing an exact TOTAL label.

    The source Service Challan sheets place the final invoice total on
    a row whose description cell contains TOTAL.
    """
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)

        has_total_label = any(
            clean_text(value).upper() == "TOTAL"
            for value in values
        )

        if not has_total_label:
            continue

        numeric_values = [
            safe_numeric(value)
            for value in values
            if isinstance(value, (int, float))
        ]

        numeric_values = [
            value
            for value in numeric_values
            if value != 0
        ]

        if numeric_values:
            return numeric_values[-1]

    return None


def choose_sc_sheet(workbook):
    """
    Select the most appropriate Service Challan sheet.
    """
    sheet_names = workbook.sheetnames

    preferred_names = [
        "TOTAL SC",
        "SERVICE CHALLAN",
        "SC",
        "SC (MEERA)",
    ]

    upper_lookup = {
        clean_text(name).upper(): name
        for name in sheet_names
    }

    for preferred_name in preferred_names:
        if preferred_name in upper_lookup:
            return upper_lookup[preferred_name]

    for sheet_name in sheet_names:
        upper_name = clean_text(sheet_name).upper()

        if (
            upper_name.startswith("SC")
            or "SERVICE CHALLAN" in upper_name
        ):
            return sheet_name

    return None


def read_standard_sc_total(file_path):
    """
    Read the final Service Challan total from a standard workbook.
    """
    workbook = load_workbook_values(file_path)

    sheet_name = choose_sc_sheet(workbook)

    if sheet_name is None:
        return None

    return find_total_in_sheet(
        workbook[sheet_name]
    )


def read_8m_sc_totals(file_path):
    """
    Read 8M totals separately for each fortnight.

    8M stores:
    SC 01-15
    SC 16-28
    """
    workbook = load_workbook_values(file_path)

    totals = {}

    for sheet_name in workbook.sheetnames:
        upper_name = clean_text(sheet_name).upper()

        if not upper_name.startswith("SC "):
            continue

        fortnight = extract_fortnight(
            file_path,
            sheet_name,
        )

        total = find_total_in_sheet(
            workbook[sheet_name]
        )

        if total is not None:
            totals[fortnight] = total

    return totals


def build_reconciliation_row(
    file_path,
    fortnight,
    source_total,
    parsed_total,
):
    """Create one reconciliation result row."""
    if source_total is None:
        difference = None
        status = "TOTAL NOT FOUND"

    else:
        difference = parsed_total - source_total

        if abs(difference) <= RECONCILIATION_TOLERANCE:
            status = "MATCH"

        else:
            status = "MISMATCH"

    return {
        "Direction": extract_direction(file_path),
        "Airline": extract_airline(file_path),
        "Month": extract_month(file_path),
        "Fortnight": fortnight,
        "Source File": Path(file_path).name,
        "Parsed Total": round(parsed_total, 4),
        "Source SC Total": (
            round(source_total, 4)
            if source_total is not None
            else None
        ),
        "Difference": (
            round(difference, 4)
            if difference is not None
            else None
        ),
        "Status": status,
    }


def reconcile_file(file_path, consolidated_df):
    """
    Reconcile one source workbook against parsed AWB totals.
    """
    file_path = Path(file_path)

    source_rows = consolidated_df[
        consolidated_df["Source File"]
        == file_path.name
    ].copy()

    airline = extract_airline(file_path)

    results = []

    if airline == "8M":
        source_totals = read_8m_sc_totals(
            file_path
        )

        for fortnight, source_total in source_totals.items():
            fortnight_rows = source_rows[
                source_rows["Fortnight"]
                == fortnight
            ]

            parsed_total = safe_numeric(
                fortnight_rows[
                    "Total Amt (Incl GST)"
                ].sum()
            )

            results.append(
                build_reconciliation_row(
                    file_path=file_path,
                    fortnight=fortnight,
                    source_total=source_total,
                    parsed_total=parsed_total,
                )
            )

        return results

    if airline == "BZ":
        summary = read_bz_summary(file_path)

        source_total = summary.get(
            "grand_total",
            0.0,
        )

    else:
        source_total = read_standard_sc_total(
            file_path
        )

    parsed_total = safe_numeric(
        source_rows[
            "Total Amt (Incl GST)"
        ].sum()
    )

    results.append(
        build_reconciliation_row(
            file_path=file_path,
            fortnight=extract_fortnight(file_path),
            source_total=source_total,
            parsed_total=parsed_total,
        )
    )

    return results


def reconcile_all(
    source_files,
    consolidated_df,
    skipped_reference_files=None,
):
    """
    Reconcile all parsed billing source workbooks.

    Reference/master workbooks are intentionally excluded.
    """
    skipped_reference_files = set(
        str(Path(path))
        for path in (skipped_reference_files or [])
    )

    results = []

    for file_path in source_files:
        normalised_path = str(Path(file_path))

        if normalised_path in skipped_reference_files:
            continue

        file_results = reconcile_file(
            file_path=file_path,
            consolidated_df=consolidated_df,
        )

        results.extend(file_results)

    columns = [
        "Direction",
        "Airline",
        "Month",
        "Fortnight",
        "Source File",
        "Parsed Total",
        "Source SC Total",
        "Difference",
        "Status",
    ]

    return pd.DataFrame(
        results,
        columns=columns,
    )